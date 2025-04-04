import polars as pl
import os
import openpyxl
import docker

path_files = '/home/oem/Документы/Inventory/03.Март/'
files_inv = os.listdir(path_files)


def insert_inv() -> pl.DataFrame:

    wb = openpyxl.open(filename=path_files + files_inv[0])
    ws = wb['Sheet1']
    data = []
    for i in range(len(ws['O:O'])):
        if i == 10:
            continue

        elif ws['O' + str(i + 1)].value == 1: #опорный столбец для извлечения данных
            data.append([ws['C' + str(i + 1)].value,
                         ws['H' + str(i + 1)].value,
                         ws['J' + str(i + 1)].value,
                         ws['L' + str(i + 1)].value,
                         ws['P' + str(i + 1)].value,
                         ws['Q' + str(i + 1)].value])

    df = pl.DataFrame(data, schema=['Наименование, назначение и краткая характеристика объекта',
                                    'Сайт',
                                    'инвентарный',
                                    'системный',
                                    'Первоначальная стоимость, руб.коп',
                                    'Остаточная балансовая стоимость,руб.коп' ], orient="row")
    return df

print(insert_inv())
# def df_divergence() -> pl.DataFrame:


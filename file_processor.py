import polars as pl
import openpyxl
from datetime import datetime


class FileProcessor:
    @staticmethod
    def process_inventory(filepath):
        try:
            wb = openpyxl.open(filename=filepath)
            ws = wb['Sheet1']
            data = []

            for i in range(1, len(ws['O']) + 1):
                if i == 11:
                    continue

                if ws['O' + str(i)].value == 1:
                    data.append([
                        ws['C' + str(i)].value or '',
                        ws['H' + str(i)].value or '',
                        ws['J' + str(i)].value or '',
                        ws['L' + str(i)].value or '',
                        float(ws['P' + str(i)].value or 0),
                        float(ws['Q' + str(i)].value or 0)
                    ])

            df = pl.DataFrame(data, schema=[
                'Наименование, назначение и краткая характеристика объекта',
                'Сайт',
                'инвентарный',
                'системный',
                'Первоначальная стоимость, руб.коп',
                'Остаточная балансовая стоимость,руб.коп'
            ], orient="row")

            return df

        except Exception as e:
            raise ValueError(f"Ошибка обработки файла: {str(e)}")